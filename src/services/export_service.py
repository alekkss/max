"""Сервис для экспорта данных в Excel."""

from typing import List
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.repositories.user_repository import IUserRepository
from src.repositories.message_repository import IMessageRepository
from src.models.user import User
from src.models.message import Message


class ExportService:
    """Сервис для экспорта данных в Excel.
    
    Отвечает за:
    - Формирование отчётов по пользователям
    - Генерацию Excel файлов
    - Форматирование данных для экспорта
    """
    
    def __init__(
        self,
        user_repository: IUserRepository,
        message_repository: IMessageRepository
    ) -> None:
        """Инициализация сервиса.
        
        Args:
            user_repository: Репозиторий пользователей
            message_repository: Репозиторий сообщений
        """
        self._user_repository = user_repository
        self._message_repository = message_repository
    
    def export_all_users_to_excel(self, output_path: str = "users_export.xlsx") -> str:
        """Экспортировать всех пользователей и их сообщения в Excel.
        
        Args:
            output_path: Путь для сохранения файла
            
        Returns:
            Абсолютный путь к созданному файлу
        """
        workbook = Workbook()
        
        # Удаляем дефолтный лист
        workbook.remove(workbook.active)
        
        # Создаём лист с пользователями
        self._create_users_sheet(workbook)
        
        # Создаём лист с сообщениями
        self._create_messages_sheet(workbook)
        
        # Сохраняем файл
        file_path = Path(output_path)
        workbook.save(file_path)
        
        return str(file_path.absolute())
    
    def _create_users_sheet(self, workbook: Workbook) -> None:
        """Создать лист с данными пользователей."""
        sheet = workbook.create_sheet("Пользователи")
        
        # Заголовки
        headers = ["User ID", "Имя", "Телефон", "Первый контакт", "Последний контакт", "Всего сообщений", "Ответов оператора"]
        sheet.append(headers)
        
        # Форматирование заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Получаем всех пользователей (нужно добавить метод в репозиторий)
        users = self._get_all_users()
        
        # Заполняем данные
        for user in users:
            # Получаем статистику по сообщениям
            all_messages = self._message_repository.get_user_messages(user.user_id, limit=10000)
            total_messages = len(all_messages)
            operator_replies = self._message_repository.count_operator_replies(user.user_id)
            
            row = [
                user.user_id,
                user.name,
                user.phone_number or "-",
                user.first_contact.strftime("%Y-%m-%d %H:%M:%S"),
                user.last_contact.strftime("%Y-%m-%d %H:%M:%S"),
                total_messages,
                operator_replies
            ]
            sheet.append(row)
        
        # Автоширина колонок
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_messages_sheet(self, workbook: Workbook) -> None:
        """Создать лист со всеми сообщениями."""
        sheet = workbook.create_sheet("Сообщения")
        
        # Заголовки
        headers = ["ID", "User ID", "Имя пользователя", "Текст", "Направление", "Оператор", "Дата/Время"]
        sheet.append(headers)
        
        # Форматирование заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Получаем всех пользователей
        users = self._get_all_users()
        
        # Собираем все сообщения
        all_messages: List[tuple[Message, str]] = []
        
        for user in users:
            messages = self._message_repository.get_user_messages(user.user_id, limit=10000)
            for message in messages:
                all_messages.append((message, user.name))
        
        # Сортируем по времени
        all_messages.sort(key=lambda x: x[0].timestamp)
        
        # Заполняем данные
        for message, user_name in all_messages:
            direction_text = "От клиента" if message.direction.value == "from_user" else "К клиенту"
            
            row = [
                message.id,
                message.user_id,
                user_name,
                message.text,
                direction_text,
                message.operator_name or "-",
                message.timestamp.strftime("%Y-%m-%d %H:%M:%S")
            ]
            sheet.append(row)
        
        # Автоширина колонок
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 80)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _get_all_users(self) -> List[User]:
        """Получить всех пользователей из БД.
        
        Временное решение - получаем через SQL напрямую.
        В будущем нужно добавить метод get_all() в IUserRepository.
        """
        # Получаем доступ к БД через репозиторий
        with self._user_repository._db.cursor() as cursor:
            cursor.execute('''
                SELECT user_id, name, first_contact, last_contact, phone_number
                FROM users
                ORDER BY last_contact DESC
            ''')
            
            rows = cursor.fetchall()
            users = []
            
            for row in rows:
                user = User(
                    user_id=row["user_id"],
                    name=row["name"],
                    first_contact=datetime.fromisoformat(row["first_contact"]),
                    last_contact=datetime.fromisoformat(row["last_contact"]),
                    phone_number=row["phone_number"]
                )
                users.append(user)
            
            return users
